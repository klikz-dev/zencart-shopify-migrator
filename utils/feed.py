import openpyxl


def read_excel(file_path: str, column_map: dict, exclude: dict, header_id=1, get_other_attributes=True):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active

    header = []
    columns = {}
    data = []

    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i == header_id - 1:

            header = row

            # Map provided field names to columns
            for idx, col in enumerate(row):
                for field_name, mapped_col_name in column_map.items():
                    if col == mapped_col_name and field_name not in columns.values():
                        # Map index to field name and stop checking for this particular header
                        columns[idx] = field_name
                        break

            break

    for i, row in enumerate(sheet.iter_rows(min_row=header_id+1, values_only=True)):

        # Construct a dictionary for each row using the mapped columns
        row_data = {field_name: row[col_idx] for col_idx,
                    field_name in columns.items() if col_idx < len(row)}

        # Detect and add attributes not explicitly mapped
        if get_other_attributes:
            attributes = {header[idx]: str(col) for idx, col in enumerate(row)
                          if col and header[idx] not in column_map.values() and header[idx].strip() not in exclude}
            row_data['attributes'] = attributes

        data.append(row_data)

    return data
